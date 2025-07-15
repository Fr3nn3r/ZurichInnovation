# This script is a reporting tool designed to convert the JSON output from an
# evaluation process (like `evaluate_file.py`) into a single, well-formatted,
# and human-readable Excel spreadsheet.
#
# The script's main functionalities are:
# 1.  **JSON Aggregation**: It is capable of reading and parsing multiple JSON
#     evaluation files in a single run. The data from all provided files is
#     combined into one dataset.
#
# 2.  **Data Transformation**: It processes the hierarchical JSON structure into a
#     flat, tabular format suitable for a spreadsheet. Each row in the final
#     Excel file represents a single "clause," and each column represents a
#     specific "rule" that the clause was evaluated against.
#
# 3.  **Data Enrichment**: It adds a 'Filename' column to the spreadsheet, which
#     is derived from the name of the source JSON file. This provides crucial
#     context, allowing users to trace each clause back to its original document.
#
# 4.  **Column Organization**: The columns in the output Excel file are logically
#     reordered. The 'Filename' and 'Clause Text' columns are placed first for
#     easy identification, followed by the rule columns, which are sorted
#     alphabetically.
#
# 5.  **Conditional Formatting**: The script applies conditional formatting to the
#     cells containing scores, making the report easier to interpret at a glance:
#     - Scores above 90 are highlighted in green.
#     - Scores between 70 and 90 are highlighted in yellow.
#     This visual cue helps to quickly identify high-confidence matches and areas
#     that may require further review.
#
# 6.  **Excel Styling**: Beyond cell coloring, the script also automatically adjusts
#     the width of each column to fit the content, ensuring that the final report
#     is clean and readable without manual adjustments.
#
# 7.  **Command-Line Interface**: The script is run from the command line and accepts
#     one or more paths to the JSON evaluation files as arguments.
#
# Usage:
#   python src/export_to_excel.py <path_to_eval1.json> [<path_to_eval2.json> ...]
#
# Example:
#   python src/export_to_excel.py output/Case-1_evaluation.json output/Case-2_evaluation.json

import pandas as pd
import json
import argparse
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def style_cells(val):
    if isinstance(val, (int, float)):
        if val > 90:
            return "background-color: #C6EFCE"  # Green for > 90
        elif val > 70:
            return "background-color: #FFEB9C"  # Yellow for > 70
    return ""


def export_to_excel(json_paths, output_path="output/combined_report.xlsx"):
    """
    Reads a list of JSON evaluation files, combines them into a single DataFrame,
    adds a 'Filename' column, and exports to a styled Excel file.
    """
    all_data = []

    for json_path in json_paths:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Extract original filename from the json path
        base_filename = os.path.basename(json_path)
        original_filename = base_filename.replace("_evaluation.json", "")

        # Process clause evaluations
        for clause in data.get("clause_level_evaluation", []):
            row = {"Clause Text": clause["clause_text"], "Filename": original_filename}
            for score in clause["evaluations"]:
                row[score["rule_name"]] = score["score"]
            all_data.append(row)

    if not all_data:
        print("No data to export.")
        return

    df = pd.DataFrame(all_data)

    # Reorder columns to have Filename first
    rule_names = [col for col in df.columns if col not in ["Clause Text", "Filename"]]
    new_column_order = ["Filename", "Clause Text"] + sorted(rule_names)
    df = df[new_column_order]

    # Fill NaN values for cleaner output
    df.fillna("-", inplace=True)

    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Clause Evaluation Report"

    # Add headers
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)

    # Apply styling
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    yellow_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )

    header_row = [cell.value for cell in ws[1]]
    rule_indices = {rule: i for i, rule in enumerate(header_row) if rule in rule_names}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for rule, col_idx in rule_indices.items():
            cell = row[col_idx]
            if isinstance(cell.value, (int, float)):
                if cell.value > 90:
                    cell.fill = green_fill
                elif cell.value > 70:
                    cell.fill = yellow_fill

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(output_path)
    print(f"Combined Excel report saved to: {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Export evaluation JSON to a styled Excel file."
    )
    parser.add_argument(
        "json_paths", nargs="+", help="One or more paths to evaluation JSON files."
    )
    args = parser.parse_args()

    export_to_excel(args.json_paths)
